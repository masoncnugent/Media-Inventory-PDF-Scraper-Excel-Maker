from PyPDF2 import PdfReader
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series_factory import SeriesFactory

#break everything into multiple python files so you don't have to scroll through the holy texts every time you want to read something

#implementing this will involve keeping the old code until this works

class PDF_Storage():

    def __init__(self):
        #once home think about a substructure of how to organize all this...
        self.pdf_list = []



class PDF:
    #class variables
    pdf_id = 1

    def __init__(self, unformatted_data, filename):
        #instance variables
        self.id = PDF.pdf_id
        #filename portion is the 0th index as a string
        self.filename = filename

        self.FTM_400S_Needed = False
        self.FTM_1000_Needed = False
        self.DFD_1000_Needed = False
        self.Remove_OD = False
        #data portion is the 1st index as a list of all the lines
        self.data = self.data_correcter(unformatted_data)
        self.inventory_list = self.inventory_list_maker()
        self.length = self.length_checker()


        PDF.pdf_id += 1

    #useful stuff for this to have...
    #the ability to index a certain line of the pdf
    #knowledge of bools like ftm 400, 1000, and dfd 1000 needed

    def data_correcter(self, unformatted_data):
        #data that is split into functional phrases without 1000 FTM and 1000 DFD added to the Excel pdf representation
        #update comment description
        spaced_data = smart_splitter(unformatted_data)

        formatted_data = data_formatter(self, spaced_data)

        return formatted_data



    #returns a particular line of the pdf
    def line_retrieval(self, line_num):
        #the '-1' turns the line_num into a list indice
        return self.data[line_num - 1]


    #gets the length of the pdf based on its self.data instance variable

    def length_checker(self):
        length = 0

        for line in self.data:
            length += 1
        
        return length
    
    def inventory_list_maker(self):
        inventory_list = []
        for pdf_line in self.data:

            try:
                #adds to the inventory list only if the line has a integer at the end
                inventory_list.append(int(pdf_line[-1]))
            
            except:
                continue
        
        return inventory_list



#changes the directory to where inventory pdfs are stored
def directory_changer():
    print("current path:")
    print(os.getcwd())

    #manual version
    #os.chdir(input("Where are the inventory files stored?\n"))

    #work version
    os.chdir(r"P:\Public\Microbiology\Media Prep\Media Inventory\2023")

    #home version
    #os.chdir(r"C:\Users\Mason\Documents\NAMSA Test\Inventory")



#helper function for smart_splitter() that gives it the read_ahead to add to phrase to make conjoined_phrase. This allows for Excel cells to have meaningful 'categories' in cells as opposed to the raw data all being in separate cells, which would ruin formatting
def read_ahead(list, r_i, read=None):

    #special case list 1 and 2 form the bulk of the cases where smart_splitter() should add the read_ahead to a given phrase
    special_case_list1 = ["o", "h", "%"]

    special_case_list2 = ["pe", "ys"]

    read = ""

    #r_i is used in place of i because it needs to progress through the list of scraped data independently from it
    r_i += 1

    #continues as long as the read index, or r_i, is less than the length of the 'list', (now just text from the scraped PDFs)
    while r_i < len(list):
        
        #creates the initial read without reading ahead for special cases that should be added onto the read by smart_splitter()
        if list[r_i] != " ":
            read += list[r_i]

            r_i += 1

        #reading ahead is done whenever a space is encountered, since it could indicate a meaningful addition to phrase
        elif list[r_i] == " ":
            if (
                #covers the two special case categories where read_ahead should be added to phrase
                read[-1] in special_case_list1
                or read[-2:] in special_case_list2

                #adds 'trays (x)' to the list of special cases (but wouldn't special_case_list2's 'ys' cover this? Is this not some other exception?)
                or len(read) > 4
                and read[-5] == "s"

                #adds Saline P80 to the list of special cases, since having '80' be a special case would cause issues
                or len(read) == 3
                and read == "P80"
                or (len(read) == 3 and read[-1] == ")")
                or (len(read) == 4 and read[-1] == ")")
            ):
                
                #recursively calls read_ahead() so long as there is a special case in front of read
                future_read = read_ahead(list, r_i, read)

                #each call will get progressively longer so long as special cases are in front of read
                if future_read != None:
                    future_read = read + " " + future_read
                    return future_read
                    
                #cuts off the read at the end of the scraped data for this line (check) 
                else:
                    return read
            #returns None if a special read is not seen 
            else:
                return None
                
    #this might not have to be written explicitly
    return None



#turns whole lines of the pdf into a list of separate strings, each of which should hold an appropriate phrase. EX: '45 trays (5)' is a phrase, as are '165(5)' and '100'
#reading ahead of a phrase is needed to determine how long each phrase should be
#each pdf that is scraped is a list containing each line of the pdf, which is also a list, hence the input parameter
def smart_splitter(list_of_lists):
    final_lists = []

    for list in list_of_lists:
        split_list = []
        #'phrase' is used as the word to designate what is read, as it could be multiple words long
        phrase = ""
        delay = 0

        #each list at this point is a line of text as one whole string, so this iterates over every character of that line string
        for i in range(len(list)):

            #causes i to not iterate over what's already been read by read_ahead()
            if delay > 0:
                delay -= 1

                #resets phrase when reading ahead has been done
                phrase = ""

            elif list[i] != " ":
                phrase += list[i]

            #read_ahead() is called when a space is encountered, as it may designate a phrase made up of multiple words
            elif list[i] == " ":
                text_ahead = read_ahead(list, i)

                #adds the text ahead of a given word, should that text be a special case. read_ahead() is recursively called so long as special cases are encountered, so the conjoined_phrase is as long as possible
                if text_ahead != None:
                    conjoined_phrase = phrase + " " + text_ahead

                    split_list.append(conjoined_phrase)

                    #delays i beyond the full extent of what's already been read
                    delay = len(text_ahead) + 1

                #simply adds phrase to split_list, should no special cases be found ahead of it
                else:
                    split_list.append(phrase)
                    phrase = ""

        #this line accounts for the final phrase, if it wasn't already part of a special case
        split_list.append(phrase)

        #adds each list to the final_lists, which is a list containing each line list with its multi-word phrases
        final_lists.append(split_list)

    return final_lists



#will add every media type to each pdf's Excel representation, even if no media of that type were listed on the pdf for the date. Takes from the data processed by smart_splitter(). Only adds 400-S FTM, 1000 FTM, and DFD as well as removing OD lines because too much more would vastly overcomplicate this.
#since so much of this is hard-coded, there is the assumption that the pdf will remain untouched in the future. In the event that it isn't, this is where changes should most likely be made first.
def data_formatter(pdf, spaced_data):
    
    #a shallow copy was needed to prevent the loop running indefinitely
    spaced_data_copy = spaced_data.copy()

    #knowing where 1000 FTM and 1000 DFD should be on an ideally formatted pdf is what is used to reference where they should be added. Sublist_count tracks what line of the pdf we should add to on the Excel representation
    sublist_count = 0
    sublist_mod = 0

    for sub_list in spaced_data_copy:
        sublist_count += 1

        #adds 400-S FTM
        if sublist_count == 12 - sublist_mod:
            if sub_list[0] == "400-S":
                sublist_count += 1
            
            else:
                pdf.FTM_400S_Needed = True
                sublist_count += 1
                sublist_mod += 1

        #adds 1000 FTM
        elif sublist_count == 15 - sublist_mod:
            if sub_list[0] == "1000":
                sublist_count += 1
                
            else:
                pdf.FTM_1000_Needed = True
                sublist_count += 1
                sublist_mod += 1
        
        #adds 1000 DFD
        elif sublist_count == 28 - sublist_mod:
            #the or condition is given because the minimum for DFD was changed on 06/26/23 from 136 to 99
            if sub_list[3] == "136" or sub_list[3] == "99":
                sublist_count += 1

            else:
                pdf.DFD_1000_Needed = True
                sublist_count += 1
                sublist_mod == 1
        
        #removes the 'OD' or 'On Demand' line from previous pdfs
        elif sublist_count == 34 - sublist_mod:
            if sub_list[0] == "OD=":
                pdf.Remove_OD = True
    

    #runs after all the other data has been added, so that the .insert() method knows where to add 1000 FTM and 1000 DFD based on where they would be in an ideally formatted pdf
    #the final value added, "", indicates that no media for this media type was recorded on this date
    if pdf.FTM_400S_Needed:
        spaced_data_copy.insert(11, ["400-S", "9 trays/lot", "4", "9 trays (2)", "0"])

    if pdf.FTM_1000_Needed:
        spaced_data_copy.insert(13, ["1000", "17-20/lot", "36", "90 (5)", "0"])

    if pdf.DFD_1000_Needed:
        spaced_data_copy.insert(25, ["DFD","1000", "34 bottles/lot", "99", "136", "0"])

    if pdf.Remove_OD:
        spaced_data_copy.pop(-1)

    return spaced_data_copy



def data_scraper():
    #list of every pdf object
    pdf_list = []
    pdf_count = 0

    #looks for pdfs in the current directory given to find media inventory pdfs
    for filename in os.listdir(os.getcwd()):
        if filename[-4:] == ".pdf":
            pdf_count += 1

            #uses imported functions from PyPDF2 to read from pdfs
            reader = PdfReader(open(filename, "rb"))

            #this can be removed in the final version
            info = reader.metadata

            #this is the raw line by line data scraped from each pdf as a list containing each line, which is also a list
            unformatted_data = []

            #every media pdf should be one page long, this could cause issues should that requirement not be met. (UPDATE)
            for i in range(0, len(reader.pages)):
                selected_page = reader.pages[i]

                raw_text = selected_page.extract_text()

                #takes each line of the pdf as a separate string, not separated by meaningful phrases
                unformatted_data += raw_text.splitlines()

            pdf = PDF(unformatted_data, filename)

            pdf_list.append(pdf)

    return pdf_list



def excel_wb_maker():

    # makes the excel workbook
    wb = Workbook()

    #sets the relevant sheet of the Excel document that will be edited
    ws = wb.active

    #gives the name for the resulting saved Excel file
    ws.title = "Inventory Analytics.xlsx"

    #this sheet stores the graphs made from the pdf data
    wb.create_sheet("Graphs")

    return wb, ws



def excel_pdf_paster(pdf_list, ws):
    #allows the rows to be placed where they need to be, and allows for space in between pasted pdfs in Excel
    pdf_offset = 1

    #moves the dates where each pdf's data is collected to start at 'I2' and continue rightward
    date_offset = 9

    #iterates through every pdf in formatted_data
    for pdf in pdf_list:
        row_count = 1

        for row in pdf.data:
            #adds to the rows in Excel without header entries like SCDB, FTM, etc.
            if len(row) < 6:
                row.insert(0, "")

            col_count = 1

            for row_data in row:
                #gives the column letter and advances it forward for the next cell in a pdf's given row
                char = get_column_letter(col_count)
                col_count += 1

                #this line is for debug purposes, but the one below it takes the data from a row on the formatted pdf data and puts it in Excel.
                row_cell = char + str(row_count + pdf_offset)
                ws[row_cell] = row_data

            row_count += 1

        pdf_offset += row_count + 2

        #adds the filename above each pdf's pasted data.
        title_cell = "A" + str(pdf_offset - 34)
        ws[title_cell] = pdf.filename

        #adds the dates that all the inventory will be compared against
        date_cell = get_column_letter(date_offset) + "2"
        ws[date_cell] = pdf.filename[:8]

        if date_cell == "CA":
            continue

        date_offset += 1



def excel_media_type_adder(ws):

    for i in range(3, 11):
        ws["H" + str(i)] = ws["A3"].value + " " + ws["B" + str(i)].value

    for i in range(11, 19):
        ws["H" + str(i)] = ws["A11"].value + " " + ws["B" + str(i)].value

    for i in range(19, 33):
        ws["H" + str(i)] = ws["A" + str(i)].value + " " + ws["B" + str(i)].value



def excel_data_mover(ws, pdf_count):
    gap = 3
    data_offset = 0

    for col in range(9, pdf_count + 9):
        col_char = get_column_letter(col)

        for row in range(3, 33):

            gap += 1

            cell_recieve = col_char + str(row)

            cell_give = "F" + str(data_offset + row)

            #conditions for ignoring weirdly formatted cells
            if ws[cell_give].value == "OD":
                continue

            #Excel graphs seem fine ignoring empty cells
            elif ws[cell_give].value == None:
                continue

            else:
                try:
                    ws[cell_recieve] = int(ws[cell_give].value)

                except:
                    ws[cell_recieve] = ws[cell_give].value
        
        data_offset += 34
        gap = data_offset


#this is still in the testing phases
def excel_graph_maker(wb, ws):
    SCDB100_chart = LineChart()
    SCDB100_chart.title = "SCDB 100 Inventory"
    SCDB100_chart.x_axis.title = "Time"
    SCDB100_chart.y_axis.title = "Inventory"

    #max_col should be taken from the data (UPDATE)

    #new implementation!
    #this would allow for making the x axis the same for every graph while iterating through the rows for the y values
    #from the axes limits and scale documentation page of openpyxl
    #made min_row = 2 just so I could test from_rows
    SCDB100_data = Reference(ws, min_col = 9, min_row = 3, max_col = 83, max_row = 3)
    SCDB100_x = Reference(ws, min_col = 9, min_row = 2, max_col = 83, max_row = 2)
    
    #I think this is series_factory and not series...
    SCDB100_s = SeriesFactory(SCDB100_data, xvalues = SCDB100_x)
    #the .series fixed this
    SCDB100_chart.series.append(SCDB100_s)

    #okay new attempt
    ##SCDB100_chart.add_data(SCDB100_data, from_rows = True)


    ##print("SCDB100_chart.series")
    ##print(SCDB100_chart.series)

    ws.add_chart(SCDB100_chart, "H34")


    return ":)"



#data_scraper calls ---> smart_splitter whose product calls ---> read_ahead

def run_program():
    directory_changer()

    #formatted_data is twice the length it should be (FIX)
    #data scraper no longer returns a pdf_count
    pdf_list = data_scraper()

    workbook, worksheet = excel_wb_maker()

    #maybe reduce what's below this to a macro function
    excel_pdf_paster(pdf_list, worksheet)
    excel_media_type_adder(worksheet)

    print(PDF.pdf_id)
    #yeahhh
    excel_data_mover(worksheet, PDF.pdf_id)

    #test
    excel_graph_maker(workbook, worksheet)

    workbook.save(worksheet.title)

    print("Excel file complete!")

run_program()

#One assumption is that the first pdf is formatted ideally, since it's media types are used to format the rest of the data into cells for use in graphs
