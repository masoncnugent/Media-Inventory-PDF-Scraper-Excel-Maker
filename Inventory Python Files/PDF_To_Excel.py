from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from PDF_Class import PDF

#utilizes openpyxl to make an Excel workbook that can hold all the pdf data and graphs
def excel_wb_maker():

    # makes the excel workbook
    wb = Workbook()

    #sets the relevant sheet of the Excel document that will be edited
    ws = wb.active

    #gives the name for the resulting saved Excel file
    ws.title = "Inventory Analytics.xlsx"

    #this sheet stores the graphs made from the pdf data
    #wb.create_sheet("Graphs")

    return wb, ws



#copies each pdfs self.data line by line on the Excel file, with a filename above each one. Includes added media types if they weren't in the original pdf
def excel_pdf_vertical_copier(ws):
    #allows the rows for each pdf to be placed where they need to be, and allows for space in between pasted pdfs in Excel
    pdf_offset = 1

    #moves the dates where each pdf's data is collected to start at 'I2' and continue rightward
    date_offset = 3

    #iterates through every pdf in formatted_data
    for pdf in PDF.pdf_list:
        row_count = 1

        for row in pdf.data:

            col_count = 1

            for row_data in row:
                #gives the column letter and advances it forward for the next cell in a pdf's given row
                #get_column_letter() is a useful function, but I later wrote my own functions that achieve the same outcome to learn how base conversion works
                char = get_column_letter(col_count)
                col_count += 1

                #this line is for debug purposes, but the one below it takes the data from a row on the formatted pdf data and puts it in Excel.
                row_cell = char + str(row_count + pdf_offset)

                #this is so the data in Excel is not inputted as a string if it can be an integer, solely to remove the green triangles in Excel.
                try:
                    ws[row_cell] = int(row_data)
                
                except:
                    ws[row_cell] = row_data

            row_count += 1

        pdf_offset += row_count + 2

        #adds the filename above each pdf's pasted data.
        title_cell = "A" + str(pdf_offset - 34)
        ws[title_cell] = pdf.filename

        #adds the dates that all the inventory will be compared against
        date_cell = "I" + str(date_offset)
        ws[date_cell] = pdf.filename[:8]

        if date_cell == "CA":
            continue

        date_offset += 1
    
    #adds the word "Date" above its respective column
    ws["I2"] = "Date"
    ws["I2"].font = Font(bold = True)



#works from the first pdf, assuming it was formatted correctly and identically to all other pdfs
#frameshift issues occur when this assumption is not met
def excel_media_type_adder(ws):

    #start at I2 and move to J2, K2, etc.
    #the first two ranges handle the cases where SCDB and FTM have to be added to the media type titles
    #there is an assumption that nothing new will be added into SCDB and FTM, built into the lengths of each section being hardcoded
    #this allows for the minimum columns to exist alongside the data for each media type
    min_offset = 0
    for i in range(3, 11):
        ws[excel_cell_shifter("G2", x_shift = i + min_offset)] = ws["A3"].value + " " + str(ws["B" + str(i)].value)
        PDF.pdf_media_type_list.append(ws["A3"].value + " " + str(ws["B" + str(i)].value))
        min_offset += 1

    for i in range(11, 19):
        ws[excel_cell_shifter("G2", x_shift = i + min_offset)] = ws["A11"].value + " " + str(ws["B" + str(i)].value)
        PDF.pdf_media_type_list.append(ws["A11"].value + " " + str(ws["B" + str(i)].value))
        min_offset += 1

    #the end is dynamically encoded to match the length of each PDF
    for i in range(19, PDF.pdf_length + 2):
        ws[excel_cell_shifter("G2", x_shift = i + min_offset)] = ws["A" + str(i)].value + " " + str(ws["B" + str(i)].value)
        PDF.pdf_media_type_list.append(ws["A" + str(i)].value + " " + str(ws["B" + str(i)].value))
        min_offset += 1

    #adds the words "Media Type" and "Minimum" above their respective column
    for i in range(len(PDF.pdf_media_type_list)):
        inv_title_col_let = excel_cell_shifter("A1", x_shift = 9 + (2 * i))
        med_title_col_let = excel_cell_shifter("A1", x_shift = 10 + (2 * i), y_shift = 1)

        ws[inv_title_col_let] = "Media Type / Inv"
        ws[med_title_col_let] = "Minimum"
        ws[inv_title_col_let].font = Font(bold = True)
        ws[med_title_col_let].font = Font(bold = True)



#copies the data from each pdf's inventory and minimum values into the table at the top of the resulting Excel file
def excel_pdf_inventory_copier(ws):

    #this is where the data starts to be printed
    row_data_offset = 3

    for pdf in PDF.pdf_list:
        col_data_offset = 10

        inv_list = pdf.inventory_list
        min_list = pdf.minimum_list

        if len(inv_list) != len(min_list):
            raise Exception("inv_list and min_list for " + str(pdf.filename) + " are not of the same length.")

        #inv_list and minimum_list should have the same length, and the lengths of either are used interchangeably for range functions
        for i in range(len(inv_list)):
            #the column letters have to be offset for the minimum to be one ahead of the column letters for the inventory
            inv_col_let = get_column_letter(col_data_offset)
            #min_col_let = get_column_letter(col_data_offset)
            ws[inv_col_let + str(row_data_offset)] = inv_list[i]

            #the row changes to allow each inventory value for a pdf to be printed
            #the += 2 accounts for the fact that the minimum values need to be pasted alongside the inventory values
            min_col_let = get_column_letter(col_data_offset + 1)
            try:
                ws[min_col_let + str(row_data_offset)] = int(min_list[i])
            except:
                ws[min_col_let + str(row_data_offset)] = ""
            col_data_offset += 2
        
        #the column changes once each pdf has their values pasted
        row_data_offset += 1



#copies the information for how much lots above the minimum exist for each pdf's media types
def excel_pdf_lots_over_min_copier(ws):

    #determines the start_point for all the other data
    lots_over_min_start_point = "I" + str(PDF.pdf_id + 3)

    month_header_loc = excel_cell_shifter(lots_over_min_start_point, y_shift = 1)
    ws[month_header_loc] = "Month"
    ws[month_header_loc].font = Font(bold = True)

    media_offset = 1
    for media_type in PDF.pdf_media_type_list:
        #prints "Media Type"
        media_type_title_cell = excel_cell_shifter(lots_over_min_start_point, x_shift = media_offset)
        ws[media_type_title_cell] = "Media Type"
        ws[media_type_title_cell].font = Font(bold = True)

        media_offset += 2
        #prints the actual media type
        media_type_proper_cell = excel_cell_shifter(media_type_title_cell, y_shift = 1)
        ws[media_type_proper_cell] = media_type


    month_offset = 1
    for month in PDF.pdf_month_list:
        #prints the months so they move down below the "Month" header
        cur_month_loc = excel_cell_shifter(month_header_loc, y_shift = month_offset)
        ws[cur_month_loc] = month
        month_offset += 1
    

        #prints the data for each month's media lot over min numbers to match alongside the greater table's position for media type inventory and minimums
        month_row_offset = 0
    for inv_over_min_month in PDF.pdf_lots_over_min:
        month_col_offset = 0
        
        for i in range(0, len(PDF.pdf_media_type_list)):

            #shifts the cell over to the right twice each time a value is pasted into Excel, and once down whenever a month's inventory over minimum numbers have been exhausted
            inv_over_min_cell = excel_cell_shifter(lots_over_min_start_point, x_shift = 1 + month_col_offset, y_shift = 2 + month_row_offset)

            ws[inv_over_min_cell] = inv_over_min_month[i]

            month_col_offset += 2
        
        month_row_offset += 1
    
    #the y_shift is compensatory for the accumulation of month_row_offset
    average_cell_header = excel_cell_shifter(inv_over_min_cell, x_shift = 2, y_shift = -month_row_offset)
    ws[average_cell_header] = "Averages"
    ws[average_cell_header].font = Font(bold = True)

    for i in range(0, len(PDF.pdf_month_list)):
        average_data_cell = excel_cell_shifter(average_cell_header, y_shift = i + 1)
        ws[average_data_cell] = PDF.pdf_lots_over_min_averages[i]



#creates every graph in the resulting Excel file
def excel_graph_maker(ws):

    #starts at the eighth column, or H
    graph_col_offset = 8
    graph_row_offset = 0

    #where the min rows for the y values start in line_chart_lots_over_min
    min_row_y_start = PDF.pdf_id + 4

    #PDF.length is equivalent to the amount of different media types
    for i in range(1, PDF.pdf_length):

        #creates the line chart for lots above the minimum for each media type, for each month data was collected
        line_chart_lots_over_min = LineChart()
        #titles are determined dynamically
        line_chart_lots_over_min.title = PDF.pdf_media_type_list[i - 1] + " Lots Above Min"
        line_chart_lots_over_min.x_axis.title = "Month"
        line_chart_lots_over_min.y_axis.title = "Lots Above Min"

        #the usage of i allows for different references to be made for the x axis data
        lots_over_min_y_values = Reference(ws, min_col = 8 + (2 * i), min_row = min_row_y_start, max_col = 8 + (2 * i), max_row = min_row_y_start + len(PDF.pdf_month_list))

        #this is static and doesn't vary for each media type
        lots_over_min_x_values = Reference(ws, min_col = 9, min_row = min_row_y_start + 1, max_col = 9, max_row = min_row_y_start + len(PDF.pdf_month_list))


        #data is y values, and categories are x values
        line_chart_lots_over_min.add_data(lots_over_min_y_values, titles_from_data = True)
        line_chart_lots_over_min.set_categories(lots_over_min_x_values)


        #32 is where the graphs should start. Ideally this would be dynamically determined, but the formatting for this never changes
        graph_lots_over_min_anchor = excel_cell_shifter("A1", x_shift = graph_col_offset + 32, y_shift = PDF.pdf_id + len(PDF.pdf_month_list) + 5 + graph_row_offset)


        #this scales the axes so each graph shows approximately the same data
        line_chart_lots_over_min.y_axis.scaling.min = -3
        line_chart_lots_over_min.y_axis.scaling.max = 3

        ws.add_chart(line_chart_lots_over_min, graph_lots_over_min_anchor)

        #adds the metadata for each graph below it
        graph_metadata_adder(ws, graph_lots_over_min_anchor, i - 1)


        #creates the line chart for the inventory and minimum values for each media type, for every day media was recorded
        line_chart_inv_min = LineChart()

        #takes the title from the list of media types given as a class attribute of PDF, dynamically
        line_chart_inv_min.title = PDF.pdf_media_type_list[i - 1] + " Inventory"
        line_chart_inv_min.x_axis.title = "Date"
        line_chart_inv_min.y_axis.title = "Inventory"


        #titles_from_data allows openpyxl to determine that this range of y_values has data for the inventory and minimum series, without it being explicitly stated
        inv_min_y_values = Reference(ws, min_col = 8 + (2 * i), min_row = 2, max_col = 9 + (2 * i), max_row = PDF.pdf_id + 1)

        inv_min_x_values = Reference(ws, min_col = 9, min_row = 3, max_col = 9, max_row = PDF.pdf_id + 1)

        line_chart_inv_min.add_data(inv_min_y_values, titles_from_data = True)
        line_chart_inv_min.set_categories(inv_min_x_values)
        
        #makes each individual chart and offsets the position for the next one
        graph_inv_min_anchor = excel_cell_shifter("A1", x_shift = graph_col_offset, y_shift = PDF.pdf_id + len(PDF.pdf_month_list) + 5 + graph_row_offset)

        ws.add_chart(line_chart_inv_min, graph_inv_min_anchor)


        #starts drawing graphs lower on the screen instead of more horizontally
        graph_col_offset += 10
        if graph_col_offset > 28:
            graph_col_offset = 8

            #graphs should be offset more or less depending on how much metadata is beneath them
            #9 months is the cutoff where the metadata would otherwise conflict with the next media's metadata if this were not done
            #UPDATE this has to be tested still with over 9 months of data
            if len(PDF.pdf_month_list) > 9: 
                graph_row_offset += 23
            else:
                graph_row_offset += 20

    #the final graph with alll the lots over min data as separate series with the average data made thick, red, and smooth
    line_chart_total_lots_over_min = LineChart()
    line_chart_total_lots_over_min.title = "Total Inventory Lots Over Min"
    line_chart_total_lots_over_min.x_axis.title = "Month"
    line_chart_total_lots_over_min.y_axis.title = "Lots Above Min"

    #used to determine legitimate y_data to be used in references, so no series are made with no data, which would clutter up the key
    #play around with the range
    valid_y_data = [col_idx for col_idx in range(10, 10 + ((PDF.pdf_length) * 2), 2) if ws.cell(row = min_row_y_start, column = col_idx).value]

    y_references = [Reference(ws, min_col=col_idx, max_col=col_idx, min_row = min_row_y_start, max_row = min_row_y_start + len(PDF.pdf_month_list)) for col_idx in valid_y_data]

    #adds each reference to valid y data
    for y_ref in y_references:
        line_chart_total_lots_over_min.add_data(y_ref, titles_from_data = True)

    #the average data is the last of the series data
    average_series = line_chart_total_lots_over_min.series[-1]

    #these make its trendline thick, red, and smooth
    average_series.graphicalProperties.line.width = 150000

    average_series.graphicalProperties.line.solidFill = "FF0000"

    average_series.smooth = True

    #this is also static, as it only uses month names
    total_lots_over_min_x_values = Reference(ws, min_col = 9, min_row = PDF.pdf_id + 5, max_col = 9, max_row = PDF.pdf_id + len(PDF.pdf_month_list) + 4)

    line_chart_total_lots_over_min.set_categories(total_lots_over_min_x_values)

    #the anchor is set where another line_chart_lot_over_min would be if it were made
    graph_total_lots_over_min_anchor = excel_cell_shifter("A1", x_shift = graph_col_offset + 32, y_shift = PDF.pdf_id + len(PDF.pdf_month_list) + 5 + graph_row_offset)

    #makes this final graph larger, for legibility
    line_chart_total_lots_over_min.width = 49
    line_chart_total_lots_over_min.height = 25

    #scales the graph so the change in averages is more easy to see
    line_chart_total_lots_over_min.y_axis.scaling.min = -2
    line_chart_total_lots_over_min.y_axis.scaling.max = 4

    ws.add_chart(line_chart_total_lots_over_min, graph_total_lots_over_min_anchor)



#turns base 26 letters for Excel columns into a decimal representation, which can be added to or subtracted from easily if needed. This allows for conversion back into shifted letters by base_26_to_let
def let_to_base_10(letters, x_shift = 0):
    offset = 64
    base_26_num = 0

    for i in range(len(letters)):
        base_26_num += (ord(letters[len(letters) - i - 1]) - offset) * (26 ** i)

    return base_26_num + x_shift


#turns the decimal representation
def base_10_to_let(base_26_num):
    offset = 64
    quotient = base_26_num
    letters = ""
    while quotient > 0:
        quotient, remainder = divmod(quotient - 1, 26)

        letters = chr(remainder + offset + 1) + letters

    return letters



#returns a cell shifted up, down, left, or right, according to the shift parameters
#+3 in x_shift moves the column 3 to the right, while -2 in y_shift moves the row 2 down.
def excel_cell_shifter(cell, x_shift = 0, y_shift = 0):

    #determines where the numbers start in the cell code, considers cases like "A2" as well as "AA2", or even "AAAAA2"
    num_start_ind = -1
    letters = ""
    for char in cell:
        num_start_ind += 1
        try:
            int(char)
            break

        except:
            letters += char
            continue
    
    numbers = int(cell[num_start_ind:])

    if y_shift != 0:
        numbers = numbers + y_shift

    #runs the functions needed to advance the letters
    if x_shift != 0:
        #shifting letters is not easy, but entirely feasible by converting to a base 10 number code, assuming Excel column letters are a base 26 system
        letters = base_10_to_let(let_to_base_10(letters, x_shift))

    return letters + str(numbers)




#makes the metadata that appears below a graph for inventory over min data by month for that given media type
#factoring in graph_length and width is a bit too hard since Excel doesn't by default make graphs full cells in width and length
def graph_metadata_adder(ws, graph_anchor, media_type_indice, graph_length = None, graph_width = None):
    y_val = 0
    x_val = 0

    #metadata starts below each graph, and from the top left anchor of a given graph the first available cell below the graph is 13 cells down
    metadata_start_point = excel_cell_shifter(graph_anchor, x_shift = x_val, y_shift = 14 + y_val)

#iterates for as many months there are
    for i in range(0, len(PDF.pdf_month_list)):

        #these use excel_cell_shifter() to determine where to be relative to the start of where metadata should be printed below each graph
        mon_loc = excel_cell_shifter(metadata_start_point, x_shift = x_val, y_shift = y_val)
        inv_min_loc = excel_cell_shifter(metadata_start_point, x_shift = x_val, y_shift = 1 + y_val)
        mon_rat_loc = excel_cell_shifter(metadata_start_point, x_shift = x_val, y_shift = 2 + y_val)

        ws[mon_loc] = PDF.pdf_month_list[i]
        ws[mon_loc].font = Font(bold = True)
        ws[inv_min_loc] = "Lots / Min"


        #this should make the percentage differences not have rounding errors
        try:
            ws[mon_rat_loc] = round(PDF.pdf_lots_over_min[i][media_type_indice], 2)
        except:
            ws[mon_rat_loc] = "NA"

        #differences between the current month's lots above min and the previous are pasted below the current months ratio and colored depending on whether the change was positive or negative
        if PDF.pdf_month_list[i] != "January":
            ratio_dif_loc = excel_cell_shifter(metadata_start_point, x_shift = x_val, y_shift = 3 + y_val)

            #this runs if there is data for both the current month and the prevous month that can be compared
            try:
                lot_num_change = round((pre_month_ratio - PDF.pdf_lots_over_min[i][media_type_indice]) * -1, 2)


                ws[ratio_dif_loc] = lot_num_change

                #the '88' parts allow the color to be more muted
                red_color = "FF8888"
                green_color = "88FF88"
                blue_color = "8888FF"

                #WAIT SUPER COOL IDEA, turn the lot_num_change into a gradient color depending on how far above or below 0 you are

                if lot_num_change < 0:
                    ws[ratio_dif_loc].fill = PatternFill(start_color = red_color, end_color = red_color, fill_type = "lightGray")

                elif lot_num_change > 0:
                    ws[ratio_dif_loc].fill = PatternFill(start_color = green_color, end_color = green_color, fill_type = "lightGray")

                else:
                    ws[ratio_dif_loc].fill = PatternFill(start_color = blue_color, end_color = blue_color, fill_type = "lightGray")

            #this would occur if one of the month ratio's involves no media recorded
            except:
                ws[ratio_dif_loc] = "NA"           

        #cur_month_ratio will be ahead of pre_month_ratio since this is declared here
        pre_month_ratio = PDF.pdf_lots_over_min[i][media_type_indice]

        x_val += 1
        #should allow for the metadata to move over if there are more than 9 months in the pdfs
        if x_val == 9:
            y_val += 4
            x_val = 0



#runs all excel related functions. data_scraper() has to be run first to create the PDF class with all it's PDF objects
def excel_batch_processor():

    wb, ws = excel_wb_maker()

    excel_pdf_vertical_copier(ws)

    excel_media_type_adder(ws)

    excel_pdf_inventory_copier(ws)

    excel_pdf_lots_over_min_copier(ws)

    #could switch worksheets for this by this point

    excel_graph_maker(ws)

    wb.save(ws.title)

    print("\nExcel file complete!")

    #figure out how to remove the pdf files from the github repos
